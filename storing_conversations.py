import pandas as pd
from openpyxl import load_workbook
    
def store_response(i,transcription,text_sentiment,tone_sentiment,recommendation_response,objection_response):
    data = {
            "Index": i,
            "Message": transcription,
            "Text Sentiment": text_sentiment,
            "Tone Sentiment": tone_sentiment,
            "recommendation": recommendation_response,
            "response": objection_response
           }
    
    new_data = pd.DataFrame(data)

    file_name = "Conversation_data.xlsx"

    try:
        book = load_workbook(file_name)
        with pd.ExcelWriter(file_name,engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            new_data.to_excel(writer, index=False,header=False, startrow=book.active.max_row)

    except FileNotFoundError:
        new_data.to_excel(file_name, index=False)