import pandas as pd
from openpyxl import load_workbook
import pandas as pd
    
def store_response(transcription,text_sentiment,tone_sentiment,recommendation_response,objection_response):
    try:
        conversation_data = pd.read_excel("data/Conversation_data.xlsx")
        if not conversation_data.empty:
            new_index = conversation_data["Index"].iloc[-1] + 1 
        else:
            new_index = 0  
    except FileNotFoundError:
        new_index = 0
    data = {
            "Index": new_index,
            "Message": transcription,
            "Text Sentiment": text_sentiment,
            "Tone Sentiment": tone_sentiment,
            "recommendation": recommendation_response,
            "response": objection_response
           }
    
    new_data = pd.DataFrame(data)

    file_name = "data/Conversation_data.xlsx"

    try:
        book = load_workbook(file_name)
        with pd.ExcelWriter(file_name,engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            new_data.to_excel(writer, index=False,header=False, startrow=book.active.max_row)

    except FileNotFoundError:
        new_data.to_excel(file_name, index=False)