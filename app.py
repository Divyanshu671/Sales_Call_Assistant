import streamlit as st
from utils.audio_processing import record_audio, transcribe_audio
from utils.sentiment_analyzing import sentiment_analysis
from utils.storing_conversations import store_response
from utils.crmd_system import workflow, summary
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from pathlib import Path
import base64
import plotly.graph_objects as go
import plotly.io as pio
import tempfile
from fpdf import FPDF
from io import BytesIO
import os
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

st.markdown(
    """
    <style>
    /* App Background */
    .main {
        background: linear-gradient(135deg, #1f3c88, #1b1f3b); /* Glamorous gradient */
        color: white; /* White text for contrast */
        font-family: 'Segoe UI', sans-serif; /* Modern font */
        padding: 20px;
    }

    /* Sidebar Styling */
    [data-testid="stSidebar"] {
        background: linear-gradient(135deg, #1b1f3b, #2c3e50); /* Gradient sidebar */
        color: white;
        box-shadow: 2px 0 10px rgba(0, 0, 0, 0.5); /* Subtle shadow */
    }

    /* Sidebar titles and menu text */
    .css-1d391kg p, .css-1v3fvcr h1, .css-1v3fvcr h2 {
        color: white; /* White text in sidebar */
    }

    /* Buttons */
    .stButton button {
        background: linear-gradient(90deg, #6a11cb, #2575fc); /* Gradient button */
        color: white;
        font-size: 16px;
        font-weight: bold;
        border: none;
        border-radius: 20px; /* Fully rounded buttons */
        padding: 10px 20px;
        transition: 0.3s ease-in-out;
    }
    .stButton button:hover {
        transform: scale(1.05); /* Slight zoom effect */
        box-shadow: 0px 4px 15px rgba(0, 0, 0, 0.3); /* Glow on hover */
    }

    /* Curved box for displaying results */
    .curved-box {
        border-radius: 15px;
        background: rgba(255, 255, 255, 0.1); /* Transparent background */
        backdrop-filter: blur(10px); /* Blur effect for a glassy look */
        padding: 20px;
        margin: 20px 0;
        color: white;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.5); /* Subtle shadow */
    }

    /* Titles and subtitles */
    .title {
        font-size: 36px;
        font-weight: bold;
        text-align: center;
        background: linear-gradient(90deg, #6a11cb, #2575fc); /* Gradient title */
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .subtitle {
        font-size: 15px;
        font-weight: 250;
        text-align: center;
        color: #d1d9e6; /* Light gray text for subtitles */
    }

    /* Output text styling */
    .output-text {
        color: lightgreen; /* Vibrant teal for outputs */
        font-size: 18px;
        font-weight: bold;
    }

    /* Scrollable areas */
    .scrollable-box {
        max-height: 400px;
        overflow-y: auto;
        background: rgba(255, 255, 255, 0.05); /* Slightly transparent background */
        padding: 50px;
        border-radius: 10px;
        box-shadow: inset 0 4px 15px rgba(0, 0, 0, 0.3);
    }

    /* Positioning logos at the top-right corner */
        .logo-container {
            position: absolute;
            top: 10px;
            right: 20px;
            display: flex;
            gap: 10px;
        }
        .logo-container img {
            width: 50px;
            height: auto;
            border-radius: 10px; /* Optional rounded corners */
            box-shadow: 0 4px 8px rgba(0, 0, 0, 0.2); /* Optional shadow */
            transition: transform 0.3s ease-in-out;
        }
        .logo-container img:hover {
            transform: scale(1.1); /* Slight zoom effect on hover */
        }
        .complete {
            background-color: white;
            color: black;
            font-size: 18px;
            font-weight: bold;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

# LOGO
def encode_image_base64(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode("utf-8")
MOBILE_LOGO_PATH = "data/mobile.jpg"
LAPTOP_LOGO_PATH = "data/laptop.jpg"
if Path(MOBILE_LOGO_PATH).is_file() or not Path(LAPTOP_LOGO_PATH).is_file():
    # Encode images
    mobile_logo_base64 = encode_image_base64(MOBILE_LOGO_PATH)
    laptop_logo_base64 = encode_image_base64(LAPTOP_LOGO_PATH)
    st.markdown(
        f"""
        <div class="logo-container">
            <img src="data:image/png;base64,{mobile_logo_base64}" alt="Mobile Logo">
            <img src="data:image/png;base64,{laptop_logo_base64}" alt="Laptop Logo">
        </div>
        """,
        unsafe_allow_html=True,
    )

# Sidebar
st.sidebar.title("Features")
menu = st.sidebar.radio("Go to", ["Home", "Dashboard"])
st.markdown("<h1 class='title'>AI Sales Call Assistant</h1>", unsafe_allow_html=True)
st.markdown("<p class='subtitle'>Your AI-powered assistant for smarter sales</p>", unsafe_allow_html=True)


# Processing
if "first" not in st.session_state:
    st.session_state.first = True
if "recording_flag" not in st.session_state:
    st.session_state.recording_flag = False
if "products" not in st.session_state:
    st.session_state.products = {}
if "products_history" not in st.session_state:
    st.session_state.products_history = {}
if "final_sentiment" not in st.session_state:
    st.session_state.final_sentiment = {}


history_file = Path("data/Conversation_data.xlsx")
st.session_state.conversation_history_df = pd.read_excel(history_file, engine='openpyxl')


if len(st.session_state.products):
    for _, row in st.session_state.products.iterrows():
        brand = row.get("Brand", row.get("Company", "Unknown Brand"))
        product = row.get("Model",row.get("Product", "Unknown Product"))
        key = f"{brand} {product}"
        if key not in st.session_state.products_history:
            st.session_state.products_history[key] = 10
        else:
            st.session_state.products_history[key] += 10
    st.session_state.products = {}

def plot_sentiment_data():
    if "conversation_history_df" in st.session_state and not st.session_state.conversation_history_df.empty:
        data = st.session_state.conversation_history_df

        sentiment_map = {"Positive": 1, "Neutral": 0, "Negative": -1, "Energetic": 1, "Moderate": 0, "Calm": -1}
        data["Tone Sentiment (Numeric)"] = data["Tone Sentiment"].map(sentiment_map)
        data["Text Sentiment (Numeric)"] = data["Text Sentiment"].map(sentiment_map)

        if data[["Tone Sentiment (Numeric)", "Text Sentiment (Numeric)"]].isnull().all().all():
            st.info("No valid sentiment data available for plotting.")
        else:
            fig = go.Figure()

            fig.add_trace(go.Scatter(
                x=data["Index"].astype(str),
                y=data["Text Sentiment (Numeric)"],
                mode="lines+markers",
                name="Text Sentiment",
                line=dict(color='rgba(52, 152, 219, 0.8)', width=2),
                marker=dict(size=8, symbol="circle", color='rgba(52, 152, 219, 1)')
            ))

            fig.add_trace(go.Scatter(
                x=data["Index"].astype(str),
                y=data["Tone Sentiment (Numeric)"],
                mode="lines+markers",
                name="Tone Sentiment",
                line=dict(color='rgba(230, 126, 34, 0.8)', width=2),
                marker=dict(size=8, symbol="square", color='rgba(230, 126, 34, 1)')
            ))

            fig.update_layout(
                title="Sentiment Shifts",
                title_font=dict(family="Segoe UI, sans-serif", size=24, color="white"),
                xaxis=dict(
                    title="Index",
                    tickmode='array',
                    tickvals=np.arange(len(data)),
                    ticktext=data["Index"].astype(str),
                    showgrid=False,
                    title_font=dict(family="Segoe UI, sans-serif", size=18, color='white'),
                    tickfont=dict(family="Segoe UI, sans-serif", size=14, color='white')
                ),
                yaxis=dict(
                    title="Sentiment Value",
                    showgrid=True,
                    gridcolor='rgba(255, 255, 255, 0.1)',
                    title_font=dict(family="Segoe UI, sans-serif", size=18, color='white'),
                    tickfont=dict(family="Segoe UI, sans-serif", size=14, color='white')
                ),
                plot_bgcolor='rgb(32, 32, 32)',
                paper_bgcolor='rgb(32, 32, 32)',
                legend=dict(
                    x=0.8, y=1, traceorder='normal',
                    font=dict(family="Segoe UI, sans-serif", size=14, color='white'),
                    bgcolor='rgba(0, 0, 0, 0.5)', bordercolor='white', borderwidth=1
                ),
                margin=dict(l=50, r=50, t=50, b=50)
            )

            return fig

    else:
        st.info("No conversation history available to plot.")

def plot_process_usage():
    labels = ['Audio Recording', 'Audio Transcription', 'Tone Sentiment Analysis', 
              'Text Sentiment Analysis', 'Recommendation and Responding']
    sizes = [20, 20, 10, 10, 40]
    colors = ['#4A90E2', '#50E3C2', '#B8E986', '#E4E4E4', '#9B9B9B']

    fig = go.Figure(data=[go.Pie(labels=labels, values=sizes, hole=0.3, textinfo='none', 
                                 marker=dict(colors=colors))])

    fig.update_layout(
        title_text="Process Usage Breakdown",
        title_font=dict(family="Segoe UI, sans-serif", size=24, color='white'),
        font=dict(family="Segoe UI, sans-serif", size=14, color='white'),
        paper_bgcolor='#333333',
        plot_bgcolor='#333333', 
        annotations=[dict(text='Process Usage', x=0.5, y=0.5, font_size=15, font_family='Segoe UI', showarrow=False)],
        showlegend=True,  
        margin=dict(l=0, r=0, t=50, b=0), 
        xaxis=dict(showgrid=False, zeroline=False),
        yaxis=dict(showgrid=False, zeroline=False)
    )
    return fig

def load_conversation_data():
    try:
        history_file = Path("data/Conversation_data.xlsx")
        if history_file.is_file():
            df = pd.read_excel(history_file)
            return df
        else:
            st.info("No file found for conversation history.")
            return pd.DataFrame()
    except Exception as e:
        st.error(f"Error loading conversation data: {e}")

def color_sentiment(val):
    if val == 'positive':
        color = 'background-color: green; color: white;'  
    elif val == 'neutral':
        color = 'background-color: yellow; color: black;'  
    elif val == 'negative':
        color = 'background-color: red; color: white;'  
    else:
        color = ''  
    return color

def display_sentiment_shifts():
    df = load_conversation_data()
    
    if not df.empty:
        sentiment_df = df[['Message', 'Text Sentiment', 'Tone Sentiment']]
        
        sentiment_df = sentiment_df.style.applymap(color_sentiment, subset=['Text Sentiment', 'Tone Sentiment'])
        
        return sentiment_df

def plot_products(products_history=st.session_state.products_history):
    products = list(products_history.keys())
    values = list(products_history.values())

    colors = ["#4F6BED", "#3AA0FF", "#8A3FFC", "#0078D4", "#38A9DB"]

    fig = go.Figure(
        data=[
            go.Bar(
                x=products,
                y=values,
                marker_color=colors * (len(products) // len(colors) + 1),
                text=values,
                textposition='auto'
            )
        ]
    )

    fig.update_layout(
        plot_bgcolor="#1E1E1E",
        paper_bgcolor="#1E1E1E",
        font_color="#FFFFFF",
        title="Most Recommended",
        xaxis=dict(title="Products", tickangle=-45),
        yaxis=dict(title="Recommendation")
    )

    return fig

def generate_pdf():
    try:
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        pdf.set_font("Arial", style="B", size=16)
        pdf.cell(200, 10, txt="Dashboard Summary", ln=True, align="C")
        pdf.ln(10)

        pdf.set_font("Arial", size=12)
        pdf.multi_cell(0, 10, summary())

        pdf.add_page()
        pdf.set_font("Arial", style="B", size=14)
        pdf.cell(200, 10, txt="Sentiment Shifts Table", ln=True, align="C")
        pdf.ln(10)
        sentiment_df = display_sentiment_shifts()
        if isinstance(sentiment_df, pd.io.formats.style.Styler):
            sentiment_df = sentiment_df.data

        pdf.set_font("Arial", "B", 12) 
        pdf.cell(70, 10, "Message", border=1, align="C")
        pdf.cell(40, 10, "Text Sentiment", border=1, align="C")
        pdf.cell(40, 10, "Tone Sentiment", border=1, align="C")
        pdf.ln() 

        pdf.set_font("Arial", "", 12)
        
        for _, row in sentiment_df.iterrows():
            message = row['Message'][:50] + "..." if len(row['Message']) > 50 else row['Message'] 
            pdf.cell(70, 10, message, border=1)
            pdf.cell(40, 10, row['Text Sentiment'], border=1, align="C")  
            pdf.cell(40, 10, row['Tone Sentiment'], border=1, align="C")  
            pdf.ln()


        temp_files = []
        try:
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_image_file1, \
                 tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_image_file2:
                
                sentiment_data_image = pio.to_image(plot_sentiment_data(), format="png")
                temp_image_file1.write(sentiment_data_image)
                temp_files.append(temp_image_file1.name)
                
                process_usage_image = pio.to_image(plot_process_usage(), format="png")
                temp_image_file2.write(process_usage_image)
                temp_files.append(temp_image_file2.name)

            pdf.add_page()
            pdf.set_font("Arial", style="B", size=14)
            pdf.cell(200, 10, txt="Sentiment Analysis Chart", ln=True, align="C")
            pdf.ln(10)
            pdf.image(temp_files[0], x=10, y=None, w=190)

            pdf.add_page()
            pdf.set_font("Arial", style="B", size=14)
            pdf.cell(200, 10, txt="Process Usage Breakdown", ln=True, align="C")
            pdf.ln(10)
            pdf.image(temp_files[1], x=10, y=None, w=190)

        finally:
            for temp_file in temp_files:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
        

        pdf.add_page()
        pdf.set_font("Arial", style="B", size=14)
        pdf.cell(200, 10, txt="Most Recommended Products", ln=True, align="C")
        pdf.ln(10)

        product_plot_image = pio.to_image(plot_products(), format="png")
        
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as temp_image_file:
            temp_image_file.write(product_plot_image)
            temp_image_file_path = temp_image_file.name

        pdf.image(temp_image_file_path, x=10, y=None, w=190)
        os.remove(temp_image_file_path)

        pdf_buffer = BytesIO()
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf_file:
            temp_pdf_path = temp_pdf_file.name  
            pdf.output(temp_pdf_path)
            
        with open(temp_pdf_path, "rb") as f:
            pdf_buffer.write(f.read())

        pdf_buffer.seek(0) 
        
        os.remove(temp_pdf_path)

        return pdf_buffer.getvalue()

    except Exception as e:
        st.error(f"Error while generating PDF: {e}")
        return None

###############################################################################################################################

if menu == "Home":
    st.markdown("#### Ask about products by telling me")
    col1, col2 = st.columns([2, 1])
    output_area = st.container()

    def process_audio_and_analyze():
        while True:
            if not st.session_state.recording_flag:
                break
            try:
                with st.spinner("I am listening..."):
                    record_audio()
                with st.spinner("Processing audio..."):
                    transcription = transcribe_audio()
                    transcription = transcription if len(transcription) else "N"
                    st.session_state.final_sentiment = sentiment_result = sentiment_analysis("data/output.wav", transcription)
                    query = [transcription, sentiment_result["Text"], sentiment_result["Tone"]]
                    response = workflow(query)
                    st.session_state.products = response[0]
                    store_response(
                    [query[0]], [query[1][0]], [query[2]["tone"]], [response[1]], [response[2]]
                    )
                    content = f"""<div class='curved-box scrollable-box'>
                        Query:{query[0]}
                        \nText Sentiment:{query[1][0]}
                        \nTone Sentiment:{query[2]["tone"]}
                        \nRecommendation:{response[1]}
                        \n<div class='output-text'>Response:{response[2]}</div>
                    </div>"""
                    with output_area:
                        st.markdown(content,unsafe_allow_html=True)
            except Exception as e:
                st.write(e)
                break
    
    with col1:
        if st.button("Say it to me 🎤"):
            st.info("Recording... ")
            st.session_state.recording_flag = True
            process_audio_and_analyze()
    with col2:
        if st.button("Stop"):
            st.session_state.recording_flag = False
            st.info("Recording Stopped!")
            with output_area:
                st.empty()
    if len(st.session_state.conversation_history_df):
        combined_content = f"""<div class='curved-box scrollable-box'>"""
        st.markdown("<h3 class='subtitle'>AI Assistant:</h3>", unsafe_allow_html= True)
        for i, row in st.session_state.conversation_history_df.iterrows():
            content = f"""<div class='curved-box scrollable-box'>
                Query:{row["Message"]}
                \nText Sentiment:{row['Text Sentiment']}
                \nTone Sentiment:{row['Tone Sentiment']}
                \nRecommendation:{row['recommendation']}
                \n<div class='output-text'>Response:{row['response']}</div>
            </div>"""
            combined_content+=content
        combined_content += """</div>"""
        st.markdown(combined_content,unsafe_allow_html=True)

elif menu == "Dashboard":
    st.markdown("<h2 class='title'>Summary</h2>", unsafe_allow_html= True)
    if not len(st.session_state.conversation_history_df):
        st.info("No Coversation History Found.")
    else:
        col1, col2 = st.columns([2, 1])
        with col1:
            if st.button("Delete Conversations"):
                try:
                    history_file = Path("data/Conversation_data.xlsx")
                    if history_file.is_file():
                        workbook = load_workbook(history_file)
                        sheet = workbook.active
                        sheet.delete_rows(2, sheet.max_row + 1)  
                        workbook.save(history_file)  
                        workbook.close()
                        
                        st.success("All conversations deleted successfully!")
                        st.session_state.conversation_history_df = pd.DataFrame(columns=["Message", "Text Sentiment", "Tone Sentiment", "recommendation", "response"])
                    else:
                        st.info("No file found to delete.")
                except Exception as e:
                    st.error(f"Error while deleting conversations: {e}")
            
        with col2:
            if len(st.session_state.conversation_history_df):
                pdf_file = generate_pdf()
                if pdf_file:
                    st.download_button(
                        label="📥 Download PDF",
                        data=pdf_file,
                        file_name="dashboard_report.pdf",
                        mime="application/pdf",
                    )
            else:
                st.info("No summary available.")
        
        # with st.expander("View Full summary", expanded=True):
        if len(st.session_state.conversation_history_df):
            summary = summary()
            sentiment_df = display_sentiment_shifts()
            sentiment_data = plot_sentiment_data()
            process_usage = plot_process_usage()

            st.markdown(f"""<div class='curved-box'>{summary}</div>""",unsafe_allow_html=True)
            st.plotly_chart(sentiment_data)
            st.markdown("### Data Metrics")
            st.dataframe(sentiment_df, use_container_width=True)
            st.plotly_chart(process_usage)
        else:
            st.info("No Conversation to plot graphs.")
        if len(st.session_state.products_history):
            products = plot_products()
            st.plotly_chart(products)
        else:
            st.info("No products Available to recommend.")